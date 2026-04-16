from dataclasses import dataclass

import pandas as pd
from openpyxl import load_workbook

import helpers.utilities as utils
from process_data.base_config import BaseConfig

# Selected projects from Uchwała ZWL Nr 787/23 (5 września 2023)
# amended by Uchwała Nr 344/24 (12 marca 2024, only dept name change for EPI04)
SELECTED_PROJECTS = {
    # Voivodeship pool
    "W16", "W20", "W18", "W02", "W12", "W08", "W15", "W03",
    # Bełchatowski
    "EBE08", "EBE06", "EBE09", "EBE01",
    # Brzeziński
    "EBR03", "EBR02", "EBR04", "EBR06", "EBR08",
    # Łaski
    "ELA14", "ELA04", "ELA06", "ELA05", "ELA01",
    # Łęczycki (ELE prefix)
    "ELE08", "ELE10", "ELE03", "ELE01",
    # Łowicki (ELC prefix)
    "ELC11", "ELC16", "ELC12", "ELC07", "ELC04",
    # Łódzki Wschodni
    "ELW05", "ELW09", "ELW01", "ELW04", "ELW10",
    # Opoczyński
    "EOP10", "EOP07", "EOP08", "EOP09", "EOP05",
    # Pabianicki
    "EPA03", "EPA02", "EPA04", "EPA09", "EPA06",
    # Pajęczański
    "EPJ09", "EPJ06", "EPJ01", "EPJ10",
    # Piotrkowski
    "EPI08", "EPI04", "EPI06", "EPI01", "EPI03",
    # Poddębicki
    "EPD06", "EPD07", "EPD05", "EPD02", "EPD10",
    # Radomszczański
    "ERA02", "ERA07", "ERA04", "ERA05", "ERA03",
    # Rawski
    "ERW06", "ERW01", "ERW10", "ERW07",
    # Sieradzki
    "ESI19", "ESI06", "ESI10", "ESI15", "ESI07",
    # Skierniewicki
    "ESK10", "ESK01", "ESK02", "ESK04",
    # Tomaszowski
    "ETM19", "ETM18", "ETM01", "ETM11", "ETM04",
    # Wieluński (EWI prefix)
    "EWI08", "EWI05", "EWI09", "EWI07",
    # Wieruszowski (EWE prefix)
    "EWE09", "EWE03", "EWE07", "EWE05",
    # Zduńskowolski
    "EZD05", "EZD18", "EZD10", "EZD15",
    # Zgierski
    "EZG04", "EZG05", "EZG02", "EZG09",
    # Łódź (miasto)
    "ELO03", "ELO07", "ELO08", "ELO06", "ELO01",
    # Skierniewice (miasto)
    "ESO03", "ESO01", "ESO07", "ESO04",
}

# Map project code prefix (3 letters after "E") → district name
# Follows car registration plate prefixes for Łódź Voivodeship
CODE_PREFIX_TO_DISTRICT = {
    "EBE": "Belchatowski",
    "EBR": "Brzezinski",
    "ELA": "Laski",
    "ELC": "Lowicki",
    "ELE": "Leczycki",
    "ELO": "Lodz",
    "ELW": "Lodzki_Wschodni",
    "EOP": "Opoczynski",
    "EPA": "Pabianicki",
    "EPD": "Poddebicki",
    "EPI": "Piotrkowski",
    "EPJ": "Pajeczanski",
    "ERA": "Radomszczanski",
    "ERW": "Rawski",
    "ESI": "Sieradzki",
    "ESK": "Skierniewicki",
    "ESO": "Skierniewice",
    "ETM": "Tomaszowski",
    "EWE": "Wieruszowski",
    "EWI": "Wielunski",
    "EZD": "Zdunowolski",
    "EZG": "Zgierski",
}

VOIVODESHIP_DISTRICT = "CITYWIDE"


def district_from_code(code):
    prefix = code[:3]
    return CODE_PREFIX_TO_DISTRICT.get(prefix)


@dataclass(kw_only=True)
class Preprocess(BaseConfig):
    preprocess: dict

    def __post_init__(self):
        self.projects_source = self.preprocess["projects_source"]
        self.votes_source = self.preprocess["votes_source"]
        self.projects_excel = self.preprocess["projects_excel"]
        self.votes_excel = self.preprocess["votes_excel"]
        self.data_dir = self.preprocess["data_dir"]
        return super().__post_init__()

    def load_workbook(self, name):
        path = utils.get_path_to_file_by_unit(
            name, self.unit, self.data_dir, ext="xlsx"
        )
        return load_workbook(path, read_only=True, data_only=True)

    def parse_projects(self, wb):
        rows = []

        # --- County pools ---
        ws = wb["pule powiatowe"]
        current_powiat_label = None
        current_district = None
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row[0] is not None and row[1] is None:
                # Header row for a new powiat
                current_powiat_label = row[0]
                current_district = None
                continue
            code = row[2]
            if not code:
                continue
            district = district_from_code(str(code))
            if district is None:
                # KUTNOWSKI and PIOTRKÓW TRYBUNALSKI MIASTO had no voting; skip
                continue
            if current_district is None:
                current_district = district
            rows.append({
                "project_id": str(code),
                "name": str(row[3]).strip() if row[3] else "",
                "cost": int(round(float(row[4]))) if row[4] else 0,
                "votes": int(row[1]) if row[1] is not None else 0,
                "selected": 1 if str(code) in SELECTED_PROJECTS else 0,
                "district": district,
            })

        # --- Voivodeship pool ---
        # Row 1: title, Row 2: header, Row 3+: data
        ws2 = wb["pula wojewódzka"]
        for row in ws2.iter_rows(min_row=3, values_only=True):
            code = row[2]
            if not code:
                continue
            rows.append({
                "project_id": str(code),
                "name": str(row[3]).strip() if row[3] else "",
                "cost": int(round(float(row[5]))) if row[5] else 0,
                "votes": int(row[1]) if row[1] is not None else 0,
                "selected": 1 if str(code) in SELECTED_PROJECTS else 0,
                "district": VOIVODESHIP_DISTRICT,
            })

        return pd.DataFrame(rows)

    def parse_votes(self, wb):
        ws = wb.active
        rows = []
        skipped = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            # (NumerKarty, Płeć, Wiek, NazwaPowiatu, KodProjWoj, Kod, ..., CzyWazny, Duplikat)
            voter_id = row[0]
            sex_raw = row[1]
            age = row[2]
            kod_woj = row[4]   # voivodeship project code
            kod = row[5]       # county project code
            valid = row[14]    # CzyWazny

            if not valid:
                skipped += 1
                continue

            # Normalise sex
            if sex_raw and str(sex_raw).lower().startswith("m"):
                sex = "M"
            elif sex_raw and str(sex_raw).lower().startswith("k"):
                sex = "F"
            else:
                sex = None

            # County vote
            if kod:
                district = district_from_code(str(kod))
                if district is None:
                    self.logger.warning(f"Unknown code prefix for county vote: {kod}")
                else:
                    rows.append({
                        "voter_id": int(voter_id),
                        "sex": sex,
                        "age": int(age) if age else None,
                        "vote": str(kod),
                        "district": district,
                    })

            # Voivodeship vote
            if kod_woj:
                rows.append({
                    "voter_id": int(voter_id),
                    "sex": sex,
                    "age": int(age) if age else None,
                    "vote": str(kod_woj),
                    "district": VOIVODESHIP_DISTRICT,
                })

        self.logger.info(f"Skipped {skipped} invalid ballots")
        df = pd.DataFrame(rows)
        df = df.sort_values(["voter_id", "district", "vote"]).reset_index(drop=True)
        return df

    def save_df(self, df, filename):
        path = utils.get_path_to_file_by_unit(
            filename, self.unit, self.data_dir, ext="xlsx"
        )
        df.to_excel(path, index=False)
        self.logger.info(f"Saved {len(df)} rows to {filename}.xlsx")

    def start(self):
        self.logger.info("Lodzkie: preprocessing projects...")
        projects_wb = self.load_workbook(self.projects_source)
        projects_df = self.parse_projects(projects_wb)
        self.save_df(projects_df, self.projects_excel)

        self.logger.info("Lodzkie: preprocessing votes...")
        votes_wb = self.load_workbook(self.votes_source)
        votes_df = self.parse_votes(votes_wb)
        self.save_df(votes_df, self.votes_excel)
