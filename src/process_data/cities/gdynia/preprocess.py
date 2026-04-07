import os
from dataclasses import dataclass

import numpy as np
import pandas as pd
import openpyxl
import xlrd

import helpers.utilities as utils
from process_data.base_config import BaseConfig


@dataclass(kw_only=True)
class Preprocess(BaseConfig):
    preprocess: dict = None

    def __post_init__(self):
        self.excel_filename = self.preprocess["excel_filename"]
        self.data_dir = self.preprocess["data_dir"]
        self.city_projects_filename = self.preprocess.get("city_projects_filename")
        self.district_projects_filename = self.preprocess.get(
            "district_projects_filename"
        )
        return super().__post_init__()

    def load_csv(self):
        self.logger.info("Preprocessing: load CSV...")
        csv_path = utils.get_path_to_file_by_unit(
            self.excel_filename, self.unit, self.data_dir, ext="csv"
        )
        encoding = "cp1250" if self.instance >= 2022 else "utf-8-sig"
        df = pd.read_csv(csv_path, delimiter=";", encoding=encoding)
        df.columns = [str(col).strip().strip('"') for col in df.columns]
        for column in df.select_dtypes(include="object").columns:
            df[column] = df[column].astype(str).str.strip().str.strip('"')
            df[column] = df[column].replace("nan", np.nan)
        return df

    def normalize_district(self, district):
        district = district.strip()
        return district.replace(" - ", "-")

    def get_project_rows(self, filename):
        csv_path = utils.get_path_to_file_by_unit(filename, self.unit, self.data_dir, "csv")
        xls_path = utils.get_path_to_file_by_unit(filename, self.unit, self.data_dir, "xls")
        xlsx_path = utils.get_path_to_file_by_unit(
            filename, self.unit, self.data_dir, "xlsx"
        )

        if os.path.exists(csv_path):
            csv_df = pd.read_csv(csv_path, delimiter=";", encoding="utf-8-sig")
            csv_df.columns = [str(col).strip().strip('"') for col in csv_df.columns]
            csv_df = csv_df.replace(np.nan, "")
            return csv_df.to_dict("records")

        if os.path.exists(xls_path):
            workbook = xlrd.open_workbook(xls_path)
            sheet = workbook.sheet_by_index(0)
            header = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
            rows = [
                {header[col]: sheet.cell_value(row_no, col) for col in range(sheet.ncols)}
                for row_no in range(1, sheet.nrows)
            ]
            return rows

        workbook = openpyxl.load_workbook(xlsx_path, read_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows_iter = sheet.iter_rows(values_only=True)
        header = list(next(rows_iter))
        rows = [dict(zip(header, row_values)) for row_values in rows_iter]
        return rows

    def get_first_present_value(self, row_data, *keys):
        for key in keys:
            if key in row_data and row_data[key] not in (None, ""):
                return row_data[key]
        return None

    def load_project_info(self, filename, default_subdistrict):
        rows = self.get_project_rows(filename)

        project_info = {}
        for row_data in rows:
            project_id = str(
                self.get_first_present_value(row_data, "ID projektu", "id_projektu", "id_wniosku")
            ).strip()
            status = str(self.get_first_present_value(row_data, "Status", "status")).strip()
            if "wycofany" in status.lower() or "negatywnie" in status.lower():
                continue

            district_value = self.get_first_present_value(row_data, "Dzielnica", "dzielnica")
            district = str(district_value or "").strip()
            if not district or district.lower().startswith("ogólnomi"):
                district = "CITYWIDE"
            else:
                district = self.normalize_district(district)

            project_type = str(
                self.get_first_present_value(row_data, "Typ projektu", "typ_projektu") or ""
            ).strip().lower()
            if district == "CITYWIDE":
                subdistrict = default_subdistrict
            elif project_type == "mały":
                subdistrict = "small"
            elif project_type == "duży":
                subdistrict = "large"
            else:
                raise RuntimeError(f"Unsupported Gdynia project type: {project_type}")
            project_info[project_id] = {
                "district": district,
                "subdistrict": subdistrict,
            }
        return project_info

    def join_selected_projects(self, row, project_columns):
        selected = [
            project_id
            for project_id in project_columns
            if str(row[project_id]).strip() in {"1", "1.0"}
        ]
        return ",".join(selected)

    def join_selected_district_projects(self, row, project_info, subdistrict):
        district = self.normalize_district(str(row["Dzielnica"]))
        if district in ("", "---"):
            return ""
        selected = [
            project_id
            for project_id, info in project_info.items()
            if info["district"] == district
            and info["subdistrict"] == subdistrict
            and str(row.get(project_id)).strip() in {"1", "1.0"}
        ]
        return ",".join(selected)

    def transform_df(self, df):
        self.logger.info("Preprocessing: transform df...")
        if self.instance >= 2021:
            project_info = {}
            project_info.update(
                self.load_project_info(self.city_projects_filename, "citywide")
            )
            project_info.update(
                self.load_project_info(self.district_projects_filename, "large")
            )

            project_columns = [col for col in df.columns if col in project_info]
            citywide_columns = [
                col
                for col in project_columns
                if project_info[col]["district"] == "CITYWIDE"
            ]

            age_column = self.get_first_present_value(
                {column: column for column in df.columns}, "Wiek", "wiek"
            )
            sex_column = self.get_first_present_value(
                {column: column for column in df.columns}, "Plec", "Płeć", "plec"
            )
            district_column = self.get_first_present_value(
                {column: column for column in df.columns}, "Dzielnica", "dzielnica"
            )
            df = df.rename(columns={district_column: "Dzielnica"})

            if self.instance == 2021:
                result = df[[age_column, sex_column, "Dzielnica"]].copy()
                result.columns = ["wiek", "plec", "dzielnica"]
                result.insert(0, "voter_id", np.arange(1, len(df) + 1))
            else:
                voter_id_column = self.get_first_present_value(
                    {column: column for column in df.columns},
                    "ID_karty",
                    "ID karty",
                )
                result = df[[voter_id_column, age_column, sex_column, "Dzielnica"]].copy()
                result.columns = ["ID_karty", "Wiek", "Plec", "Dzielnica"]
            result["citywide_vote"] = df.apply(
                lambda row: self.join_selected_projects(row, citywide_columns), axis=1
            )
            result["small_vote"] = df.apply(
                lambda row: self.join_selected_district_projects(
                    row, project_info, "small"
                ),
                axis=1,
            )
            result["large_vote"] = df.apply(
                lambda row: self.join_selected_district_projects(
                    row, project_info, "large"
                ),
                axis=1,
            )
            if self.instance == 2021:
                result["dzielnica"] = result["dzielnica"].replace(np.nan, "")
            else:
                result["Dzielnica"] = result["Dzielnica"].replace(np.nan, "")
                base_ids = result["ID_karty"].astype(int).astype(str)
                duplicated = result.groupby("ID_karty").cumcount()
                result["ID_karty"] = np.where(
                    duplicated == 0,
                    base_ids,
                    duplicated.astype(str).radd("99999") + base_ids,
                ).astype(int)
            return result

        df = df.replace(0, np.nan)
        citywide_df = df.filter(regex="OGM*", axis=1)
        district_df = df.filter(regex="^(?!.*OGM).*\/\d{4}", axis=1)

        result = df.iloc[:, :4]

        result["citywide_vote"] = (
            citywide_df.stack()
            .reset_index()
            .groupby("level_0")["level_1"]
            .transform(lambda x: ",".join(x))
        )
        result["district_vote"] = (
            district_df.stack()
            .reset_index()
            .groupby("level_0")["level_1"]
            .transform(lambda x: ",".join(x))
        )

        # to remove voter_id == 0
        result.index = np.arange(1, len(df) + 1)
        return result

    def start(self):
        self.logger.info("Running preprocessing...")
        excel_path = utils.get_path_to_file_by_unit(
            self.excel_filename, self.unit, self.data_dir, ext="xlsx"
        )
        if os.path.exists(excel_path):
            self.logger.info(
                "There is already existing Excel file, preprocessing "
                "stopped. If you want to convert anyway, "
                f"please remove xlsx {excel_path}"
            )
        else:
            self.preprocess_csv(excel_path)

    def preprocess_csv(self, excel_path):
        df = self.load_csv()
        df = self.transform_df(df)
        self.logger.info("Preprocessing: save df to Excel...")
        if self.instance >= 2021:
            df.to_excel(excel_path, index=False)
        else:
            df.to_excel(excel_path, index_label="voter_id")
