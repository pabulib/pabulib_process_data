import re
from dataclasses import dataclass

import pandas as pd
from openpyxl import load_workbook

import helpers.utilities as utils
from process_data.base_config import BaseConfig


@dataclass(kw_only=True)
class Preprocess(BaseConfig):
    preprocess: dict

    vote_pattern = re.compile(r"(\d+)\s*PKT\s*-\s*Nr\s*(\d+)", re.IGNORECASE)
    large_pool_label = 'Projekty z kategorii "Duże" - powyżej 410 000 zł'
    small_pool_label = 'Projekty z kategorii "Małe" - do 410 000 zł'
    valid_status = "ważny"

    def __post_init__(self):
        self.source_excel = self.preprocess["source_excel"]
        self.votes_source_excel = self.preprocess.get("votes_source_excel", self.source_excel)
        self.projects_excel = self.preprocess["projects_excel"]
        self.votes_excel = self.preprocess["votes_excel"]
        self.data_dir = self.preprocess["data_dir"]
        return super().__post_init__()

    def load_workbook(self, excel_name):
        path = utils.get_path_to_file_by_unit(
            excel_name, self.unit, self.data_dir, ext="xlsx"
        )
        return load_workbook(path, read_only=True, data_only=True)

    @staticmethod
    def clean_int(value):
        return int(re.sub(r"[^0-9]", "", str(value)))

    @staticmethod
    def clean_votes(value):
        match = re.search(r"\d[\d\s\xa0]*", str(value))
        if not match:
            raise RuntimeError(f"Cannot parse vote count from: {value}")
        return int(match.group(0).replace(" ", "").replace("\xa0", ""))

    def parse_projects(self, workbook):
        sheet = workbook["Lista projektów"]
        rows = []
        project_pool_mapping = {}
        current_pool = None

        for row in sheet.iter_rows(values_only=True):
            if row[0] == self.large_pool_label:
                current_pool = "municipal large"
                continue
            if row[0] == self.small_pool_label:
                current_pool = "municipal small"
                continue

            project_id = row[1]
            if not isinstance(project_id, int):
                continue

            if current_pool is None:
                raise RuntimeError(
                    f"Ruda_Slaska: missing project pool for project {project_id}."
                )

            project_pool_mapping[str(project_id)] = current_pool
            rows.append(
                {
                    "project_id": project_id,
                    "name": row[4],
                    "cost": self.clean_int(row[7]),
                    "votes": self.clean_votes(row[3]),
                    "score": int(row[2]),
                    "district": current_pool,
                    "selected": 1 if str(row[8]).strip().lower() == "tak" else 0,
                    "category": str(row[5]).strip().lower(),
                    "neighborhood": str(row[6]).strip(),
                }
            )

        return pd.DataFrame(rows), project_pool_mapping

    def parse_votes(self, workbook, project_pool_mapping):
        sheet_name = (
            "Zanonimizowana baza głosów"
            if "Zanonimizowana baza głosów" in workbook.sheetnames
            else workbook.sheetnames[0]
        )
        sheet = workbook[sheet_name]
        rows = []
        empty_votes = 0
        invalid_votes = 0

        headers = [str(value).strip() if value is not None else "" for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        header_indexes = {name: idx for idx, name in enumerate(headers)}

        voter_id_idx = header_indexes["Nr głosu"]
        voting_method_idx = header_indexes["Typ głosu"]
        vote_text_idx = header_indexes["Nr projektów"]
        sex_idx = header_indexes["Płeć"]
        status_idx = header_indexes.get("Status")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            voter_id = row[voter_id_idx]
            votes_text = row[vote_text_idx]
            if not voter_id:
                continue
            if status_idx is not None:
                status = str(row[status_idx]).strip().lower()
                if status != self.valid_status:
                    invalid_votes += 1
                    continue
            if not votes_text:
                empty_votes += 1
                continue

            matches = self.vote_pattern.findall(str(votes_text))
            if not matches:
                empty_votes += 1
                continue

            for points, project_id in matches:
                district = project_pool_mapping.get(project_id)
                if district is None:
                    raise RuntimeError(
                        f"Ruda_Slaska: project {project_id} from votes not found in project pool mapping."
                    )
                rows.append(
                    {
                        "voter_id": int(voter_id),
                        "sex": row[sex_idx],
                        "points": int(points),
                        "vote": project_id,
                        "district": district,
                        "voting_method": row[voting_method_idx],
                    }
                )

        if empty_votes:
            self.logger.warning(
                f"Ruda_Slaska: skipped {empty_votes} vote rows with empty vote data."
            )
        if invalid_votes:
            self.logger.info(
                f"Ruda_Slaska: skipped {invalid_votes} invalid ballots based on status column."
            )

        votes_df = pd.DataFrame(rows)
        votes_df = votes_df.sort_values(["voter_id", "vote"]).reset_index(drop=True)
        return votes_df

    def save_df(self, df, filename):
        output_path = utils.get_path_to_file_by_unit(
            filename, self.unit, self.data_dir, ext="xlsx"
        )
        df.to_excel(output_path, index=False)

    def start(self):
        self.logger.info("Running preprocessing...")
        projects_workbook = self.load_workbook(self.source_excel)
        votes_workbook = self.load_workbook(self.votes_source_excel)

        projects_df, project_pool_mapping = self.parse_projects(projects_workbook)
        votes_df = self.parse_votes(votes_workbook, project_pool_mapping)

        self.save_df(projects_df, self.projects_excel)
        self.save_df(votes_df, self.votes_excel)
