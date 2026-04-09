from collections import Counter
from dataclasses import dataclass

import openpyxl
import helpers.utilities as utils
from process_data.get_projects_excel import GetProjects as BaseGetProjects


@dataclass(kw_only=True)
class GetProjects(BaseGetProjects):
    preserve_official_results: bool = False
    votes_excel_filename: str = ""

    def create_district_upper_mapping(self):
        self.district_upper_district_mapping = {}
        for district in self.district_projects_mapping.keys():
            district_upper = utils.change_district_into_name(district)
            self.district_upper_district_mapping[district_upper] = district

    def load_counted_results(self):
        votes_path = utils.get_path_to_file_by_unit(
            self.votes_excel_filename, self.unit, extra_dir=self.data_dir, ext="xlsx"
        )
        workbook = openpyxl.load_workbook(votes_path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        headers = list(next(rows))
        voter_idx = headers.index("voter_id")
        vote_idx = headers.index("vote")
        points_idx = headers.index("points")

        counted_votes = Counter()
        counted_scores = Counter()
        seen_votes = set()

        for row in rows:
            voter_id = row[voter_idx]
            project_id = str(row[vote_idx])
            key = (voter_id, project_id)
            if key not in seen_votes:
                counted_votes[project_id] += 1
                seen_votes.add(key)
            counted_scores[project_id] += int(row[points_idx])

        return counted_votes, counted_scores

    def update_projects_with_counted_results(self):
        counted_votes, counted_scores = self.load_counted_results()

        for district, projects in self.projects_data_per_district.items():
            for project in projects:
                project_id = str(project["project_id"])
                project["votes"] = counted_votes[project_id]
                project["score"] = counted_scores[project_id]

    def start(self):
        self.prepare_excel_sheet()
        self.handle_columns_indexes()
        self.iterate_through_projects()
        if not self.preserve_official_results:
            self.update_projects_with_counted_results()
        self.create_district_upper_mapping()
        self.post_process()
        objects = {
            "district_projects_mapping": self.district_projects_mapping,
            "projects_data_per_district": self.projects_data_per_district,
            "project_district_mapping": self.project_district_mapping,
            "district_upper_district_mapping": self.district_upper_district_mapping,
        }
        self.save_mappings_as_jsons(objects)
