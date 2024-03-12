from dataclasses import dataclass

import pandas as pd
import xlrd

import helpers.utilities as utils
from process_data.base_config import BaseConfig


@dataclass(kw_only=True)
class Preprocess(BaseConfig):
    preprocess: dict

    def __post_init__(self):
        self.excel_to_preprcess = self.preprocess["excel_to_preprocess"]
        self.output_excel_path = self.preprocess["output_excel_path"]
        self.data_dir = self.preprocess["data_dir"]
        self.district_district_name_mapping = {"districts": {}, "subdistricts": {}}
        return super().__post_init__()

    def load_excel_sheet(self):
        from openpyxl import load_workbook

        excel_path = utils.get_path_to_file_by_unit(
            self.excel_to_preprcess, self.unit, self.data_dir, ext="xlsx"
        )

        workbook = load_workbook(excel_path, data_only=True)
        sheet = workbook.worksheets[0]
        return sheet

    def get_excel_data(self, sheet):
        data = []

        district = None
        subdistrict = None

        for row in sheet.iter_rows(sheet.min_row, sheet.max_row):
            cell = row[0]
            selected = "0"
            indexed = cell.fill.start_color.index
            tint = cell.fill.start_color.tint
            if tint == -0.249977111117893:
                district = cell.value
                if district == "PONADOSIEDLOWE":
                    district = "CITYWIDE"
                    subdistrict = "CITYWIDE"
                continue
            elif tint == -0.1499984740745262:
                subdistrict = cell.value
                continue
            elif cell.value == "Lp":
                column_names = [cell.value for cell in row]
                continue
            elif indexed == "FFFFFF00":
                selected = "1"

            row_values = [cell.value for cell in row]

            if "PROJEKTY ZWYCIĘSKIE" in row_values:
                continue

            data.append(row_values + [district, subdistrict, selected])
            project_id = row_values[3]
            self.add_district_name_mapping(project_id, district, subdistrict)

        column_names += ["district", "subdistrict", "selected"]

        return data, column_names

    def add_district_name_mapping(self, project_id, district, subdistrict):
        district_id = project_id[0]
        subdistrict_id = project_id[-2:]
        self.district_district_name_mapping["districts"][district_id] = district
        if district == "CITYWIDE":
            return
        self.district_district_name_mapping["subdistricts"][
            subdistrict_id
        ] = subdistrict

    def save_objects(self):
        objects = {
            "district_district_name_mapping": self.district_district_name_mapping,
        }
        self.save_mappings_as_jsons(objects)

    def start(self):
        self.logger.info("Running preprocessing...")
        sheet = self.load_excel_sheet()
        data, columns = self.get_excel_data(sheet)
        df = pd.DataFrame(data, columns=columns)
        print(df.head())
        self.logger.info("Preprocessing: save df to Excel...")
        output_path = utils.get_path_to_file_by_unit(
            self.output_excel_path, self.unit, self.data_dir, ext="xlsx"
        )
        df.to_excel(output_path, index=False)
        self.save_objects()
